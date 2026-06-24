import os
import json
import pandas as pd
from pathlib import Path
from dotenv import load_dotenv
import PyPDF2
import docx
import openpyxl
from src.processor.yandex_gpt_client import YandexGPTClient

load_dotenv()

class DocumentProcessor:
    """Обработчик документов с использованием YandexGPT"""
    
    def __init__(self):
        self.gpt_client = YandexGPTClient()
        self.processed_dir = os.getenv("PROCESSED_DIR", "processed_data")
        self.docs_dir = os.getenv("DOCS_DIR", "downloaded_docs")
        os.makedirs(self.processed_dir, exist_ok=True)
    
    def extract_text_from_pdf(self, file_path):
        """Извлекает текст из PDF"""
        try:
            text = ""
            with open(file_path, 'rb') as f:
                pdf_reader = PyPDF2.PdfReader(f)
                for page in pdf_reader.pages:
                    text += page.extract_text() + "\n"
            return text
        except Exception as e:
            print(f"Ошибка извлечения текста из PDF {file_path}: {e}")
            return ""
    
    def extract_text_from_docx(self, file_path):
        """Извлекает текст из DOCX"""
        try:
            doc = docx.Document(file_path)
            text = "\n".join([paragraph.text for paragraph in doc.paragraphs])
            return text
        except Exception as e:
            print(f"Ошибка извлечения текста из DOCX {file_path}: {e}")
            return ""
    
    def extract_text_from_xlsx(self, file_path):
        """Извлекает текст из XLSX"""
        try:
            wb = openpyxl.load_workbook(file_path)
            text = ""
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                text += f"\n--- Лист: {sheet} ---\n"
                for row in ws.iter_rows(values_only=True):
                    text += " | ".join([str(cell) if cell else "" for cell in row]) + "\n"
            return text
        except Exception as e:
            print(f"Ошибка извлечения текста из XLSX {file_path}: {e}")
            return ""
    
    def extract_text_from_txt(self, file_path):
        """Извлекает текст из TXT/CSV"""
        try:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                return f.read()
        except Exception as e:
            print(f"Ошибка извлечения текста из TXT {file_path}: {e}")
            return ""
    
    def extract_text(self, file_path):
        """Извлекает текст из файла в зависимости от расширения"""
        ext = Path(file_path).suffix.lower()
        
        if ext == '.pdf':
            return self.extract_text_from_pdf(file_path)
        elif ext in ['.docx', '.doc']:
            return self.extract_text_from_docx(file_path)
        elif ext in ['.xlsx', '.xls']:
            return self.extract_text_from_xlsx(file_path)
        elif ext in ['.txt', '.csv', '.xml']:
            return self.extract_text_from_txt(file_path)
        else:
            print(f"Неподдерживаемый формат: {ext}")
            return ""
    
    def determine_document_type(self, file_name, text_sample):
        """Определяет тип документа по имени и содержимому"""
        file_name_lower = file_name.lower()
        text_lower = text_sample.lower()
        
        if 'смет' in file_name_lower or 'смет' in text_lower:
            return "smeta"
        elif 'технич' in file_name_lower or 'тз' in file_name_lower or 'задани' in text_lower:
            return "tz"
        elif 'контракт' in file_name_lower or 'договор' in file_name_lower:
            return "contract"
        elif 'проект' in file_name_lower:
            return "project"
        elif 'обоснован' in file_name_lower or 'нмц' in file_name_lower:
            return "nmck"
        else:
            return "unknown"
    
    def process_single_document(self, file_path, tender_id):
        """Обрабатывает один документ"""
        print(f"   -> Обработка: {Path(file_path).name}")
        
        # Извлекаем текст
        text = self.extract_text(file_path)
        if not text or len(text.strip()) < 100:
            print(f"   -> ⚠️ Текст слишком короткий или пустой")
            return None
        
        # Определяем тип документа
        doc_type = self.determine_document_type(Path(file_path).name, text[:500])
        print(f"   -> Тип документа: {doc_type}")
        
        # Отправляем в YandexGPT
        result = self.gpt_client.extract_construction_data(text, doc_type)
        
        if "error" in result:
            print(f"   -> ❌ Ошибка YandexGPT: {result['error']}")
            return None
        
        # Добавляем метаданные
        result['document_name'] = Path(file_path).name
        result['document_type'] = doc_type
        result['tender_id'] = tender_id
        
        return result
    
    def process_tender_folder(self, tender_id):
        """Обрабатывает все документы в папке закупки"""
        tender_folder = os.path.join(self.docs_dir, tender_id)
        
        if not os.path.exists(tender_folder):
            print(f"Папка {tender_id} не найдена")
            return None
        
        print(f"\n{'='*60}")
        print(f"Обработка закупки: {tender_id}")
        print(f"{'='*60}")
        
        # Находим все документы
        files = [f for f in os.listdir(tender_folder) 
                if not f.startswith('_') and not f.startswith('.')]
        
        if not files:
            print(f"Нет документов для обработки")
            return None
        
        print(f"Найдено документов: {len(files)}")
        
        # Обрабатываем каждый документ
        results = []
        for file_name in files:
            file_path = os.path.join(tender_folder, file_name)
            result = self.process_single_document(file_path, tender_id)
            if result:
                results.append(result)
        
        if not results:
            return None
        
        # Сохраняем результаты
        output_file = os.path.join(self.processed_dir, f"{tender_id}_processed.json")
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(results, f, ensure_ascii=False, indent=2)
        
        print(f"✓ Результаты сохранены: {output_file}")
        return results
    
    def process_all_tenders(self):
        """Обрабатывает все закупки"""
        print("\n" + "="*60)
        print("НАЧАЛО ОБРАБОТКИ ВСЕХ ДОКУМЕНТОВ")
        print("="*60)
        
        # Получаем список всех закупок
        tender_ids = [d for d in os.listdir(self.docs_dir) 
                     if os.path.isdir(os.path.join(self.docs_dir, d))]
        
        print(f"Всего закупок для обработки: {len(tender_ids)}")
        
        all_results = []
        
        for tender_id in tender_ids:
            results = self.process_tender_folder(tender_id)
            if results:
                all_results.extend(results)
        
        # Сохраняем сводную таблицу
        if all_results:
            self.save_summary_csv(all_results)
        
        print("\n" + "="*60)
        print("ОБРАБОТКА ЗАВЕРШЕНА")
        print(f"Всего обработано документов: {len(all_results)}")
        print("="*60)
    
    def save_summary_csv(self, all_results):
        """Сохраняет сводную CSV таблицу"""
        summary_file = os.path.join(self.processed_dir, "summary.csv")
        
        # Преобразуем в плоскую структуру
        flat_data = []
        for result in all_results:
            row = {
                'tender_id': result.get('tender_id'),
                'document_name': result.get('document_name'),
                'document_type': result.get('document_type'),
                'facade_area_m2': result.get('work_volumes', {}).get('facade_area_m2'),
                'roof_area_m2': result.get('work_volumes', {}).get('roof_area_m2'),
                'length_m': result.get('work_volumes', {}).get('length_m'),
                'volume_m3': result.get('work_volumes', {}).get('volume_m3'),
                'work_types': ', '.join(result.get('work_types', [])),
                'floors': result.get('building_info', {}).get('floors'),
                'address': result.get('building_info', {}).get('address'),
                'region': result.get('building_info', {}).get('region'),
                'duration_days': result.get('duration_days'),
                'cost_rub': result.get('cost_rub'),
                'materials_count': len(result.get('materials', []))
            }
            flat_data.append(row)
        
        df = pd.DataFrame(flat_data)
        df.to_csv(summary_file, index=False, encoding='utf-8-sig')
        print(f"✓ Сводная таблица сохранена: {summary_file}")


if __name__ == "__main__":
    processor = DocumentProcessor()
    processor.process_all_tenders()